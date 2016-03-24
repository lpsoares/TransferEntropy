# Condigo do Leo convertido para Python
# Luciano 24 março 2016

import numpy as np      # mathmatical library
import openpyxl         # excel support

class xls:

    def __init__(self, filename, worksheet, debug = False):
        
        # open the worksheet and the specific sheet
        self.wb = openpyxl.load_workbook(filename)
        self.sheet = self.wb.get_sheet_by_name(worksheet)

        if debug:
            print("Total Rows    ",self.sheet.max_row)
            print("Total Columns ",self.sheet.max_column)
        return

    def getMatrix(self, last_column, debug = False):

        columns =  openpyxl.cell.column_index_from_string(last_column)
        if debug:
            print("last column read (",last_column,") = ",columns)

        assert (self.sheet.max_column >= columns),"A column bigger than the spreadsheet was specified!"

        # convert excel in a numpy matrix
        matrixValues = np.array([[i.value for i in j[1:columns]] for j in self.sheet.rows[1:]])

        if debug:
            print("primeira célula: ",matrixValues[0,0])
            print("última coluna: ",matrixValues[0,columns-2])
            print("última linha: ",matrixValues[self.sheet.max_row-2,0])
            print("última célula: ",matrixValues[self.sheet.max_row-2,columns-2])

        return matrixValues

    def getNames(self, last_column, debug = False):

        columns =  openpyxl.cell.column_index_from_string(last_column)
        if debug:
            print("last column read (",last_column,") = ",columns)

        assert (self.sheet.max_column >= columns),"A column bigger than the spreadsheet was specified!"

        # convert excel in a numpy matrix
        NameValues = np.array([i.value for i in self.sheet.rows[0][1:]])

        if debug:
            print("primeiro nome: ",NameValues[0])
            print("último nome: ",NameValues[columns-2])

        return NameValues

    def getTimes(self, debug = False):

        # convert excel in a numpy matrix
        TimeValues = np.array([j[0].value for j in self.sheet.rows[1:]])

        if debug:
            print("primeiro tempo: ",TimeValues[0])
            print("último tempo: ",TimeValues[self.sheet.max_row-2])

        return TimeValues

    def getValues(self, last_column, debug = False):
        return(self.getMatrix(last_column, debug),self.getNames(last_column, debug),self.getTimes(debug))
